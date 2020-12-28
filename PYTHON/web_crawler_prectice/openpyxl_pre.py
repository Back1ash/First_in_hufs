from openpyxl import Workbook

write_wb = Workbook()
write_ws = write_wb.create_sheet('생성시트')

write_ws = write_wb.active
write_ws.append([1,2,3])
write_ws.append([11,22,33])
write_ws.append([])
write_ws.append([])
write_ws.append(['하이루', 'ㅋㅋ'])
write_ws.cell(5, 5, '5행5열')
write_wb.save("숫자.xlsx")